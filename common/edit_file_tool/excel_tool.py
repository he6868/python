# -*- coding:utf-8 -*-
# @time :2023-6-3 17:44
# @Author :lemon_huahua
# @File: excel_tool.py
# 文件说明：excel工具类
import openpyxl
from openpyxl import load_workbook

class ExcelTool:

    # 读取excel
    def read_excel(self, excel_path):
        """
        从excel读取数据，未经过格式化
        :param excel_path: excel文件路径
        :return: excel数据
        """
        # 获取excel文件对象
        excel = load_workbook(excel_path)
        # 遍历excel下所有sheet
        for sheet_name in excel.sheetnames:
            # 根据（sheet页名称）获取sheet页对象
            sheet_obj = excel[sheet_name]

            # 遍历每个sheet页内的值
            for values, sheet_values in zip(sheet_obj.values, sheet_obj):

                values = list(values)  # 将获取的行用例转换为列表
                # 遍历每个sheet中的值：用于获取所在行数
                for value in sheet_values:

                    # 将当前用例的（excel路径、sheet名称、所在行）以列表形式存放至行用例末尾。后续会用到
                    values.append([excel_path, sheet_name, value.row])
                    break  # 退出本次循环
                # 按每行用例返回
                yield values

    # 打开并编辑excel（注意：处于性能考虑，编辑之后不会自动保存， 需要自行编写保存代码）
    def edit_excel(self, path, sheet_name, column, row, value):
        """
        编辑excel，不会清空原有数据
        :param path: excel文件路径
        :param sheet_name: sheet页名称
        :param column: 所在列
        :param row: 所在行
        :param value: 所需要写入的值
        :return:
        """
        # 打开需要编辑的文件
        workbook = openpyxl.load_workbook(path)  # 打开文件

        # 打开需要编辑的sheet页名称
        sheet = workbook[sheet_name]  # 读取sheet页

        # 需要编辑的行column、列roe，以及编辑写入的内容value
        sheet.cell(column=column, row=row, value=value)