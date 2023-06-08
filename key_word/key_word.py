# -*- coding:utf-8 -*-
# @time :2023-5-10 22:14
# @Author :lemon_huahua
# @File: key_word.py
# 文件说明：关键字类
import ast
import re
import openpyxl

from common.base import Base
from common.url_correlati.request_operate import Request
from common.edit_file_tool.yaml_tool import YamlOperate
from common.url_correlati.assertion import Assertion
from common.web_ui_correlati.selenium_operate import SeleniumTool


class Key:

    def __init__(self):
        # 文件路径定义
        self.global_path = f'{Base().get_project_path()}/data/config_data/global_variable.yaml'  # 全局文件路径
        self.config_path = f'{Base().get_project_path()}/data/config_data/config_data.yaml'  # 配置文件

    # 【接口】：--------------------------------------------------------------------------------------------------------
    # post: post请求
    def post(self,case_data):
        res_data = Request().all_request(case_data)  # 执行post请求
        # Assertion().extractor(case_data, res_data)  # 提取器
        # self.all_assertion(case_data, res_data)  # 总断言
        # return res_data

    # get: get请求
    def get(self,case_data):
        # 执行post请求
        res_data = self.post(case_data)
        return res_data

    # if判断
    def if_jud(self, case_data):
        pass

    # （URL编码）互转
    def scuc(self, case_data):
        """
        全文为：str_conversion_url_coding（字符串转url编码），并写入全局变量文件
        :return:
        """
        code_mode = str(case_data[7]).split('\n')  # 获取（编码模式），并用“换行符”分隔
        string = str(case_data[8]).split('\n')  # 获取（编码字符串），并用“换行符”分开
        mode_ = True

        # 遍历（编码模式）和（编码字符串）
        for como, str_ in zip(code_mode, string):
            # 将数据从特殊格式列表中提取出来
            str_s = Base().special_format_data_extract(str_)

            if como in ['decode', '解码']:
                # 【解码】
                mode_ = False
            url_code = Base().url_coding_conversion(str_s['front'], mode_)  # 调用编码函数
            # 将编码或解码后的数据写入（全局变量）
            YamlOperate().updata_yaml(self.global_path, {f'{str_s["after"]}': f'{url_code}'})

    # （gb2312）编码互转
    def gb2312(self, case_data):

        code_mode = str(case_data[7]).split('\n')  # 获取（编码模式），并用“换行符”分隔
        string = str(case_data[8]).split('\n')  # 获取（编码字符串），并用“换行符”分开
        mode_ = True
        # 遍历（编码模式）和（编码字符串）
        for como, str_ in zip(code_mode, string):
            # 将数据从特殊格式列表中提取出来
            str_s = Base().special_format_data_extract(str_)

            if como in ['decode', '解码']:
                # 【解码】
                mode_ = False
            url_code = Base().gb2312_code_conversion(str_s['front'], mode_)  # 调用编码函数
            # 将编码或解码后的数据写入（全局变量）
            YamlOperate().updata_yaml(self.global_path, {f'{str_s["after"]}': f'{url_code}'})

    # （全局变量）替换
    def global_var_replace(self, string):
        """
        将格式化后的数据进行全局变量替换，将用例里面的关键词进行替换，替换关键词：${变量名}
        :param case_data: 未替换变量后的用例
        :return: 替换变量后的用例
        """
        case_str = str(string)  # 将用例转换成str格式，否则无法使用正则表达式和替换
        yaml_datas = YamlOperate().read_yaml(f'{Base().get_project_path()}/data/config_data/global_variable.yaml')

        # 将关键词识别出来，并添加到列表中
        keys_list = re.findall(r'\${(.*?)}', case_str)

        # 如果关键词列表不为空，则进入（全局变量）yaml文件中寻找值并替换
        if keys_list:
            for key_var in keys_list:  # 遍历关键词列表，获取关键值
                var_ = yaml_datas[f'{key_var}']  # 提取yaml文件中的变量值
                key_words = f'${{{key_var}}}'  # 拼接关键词，将关键词加上${}
                case_str = case_str.replace(key_words, f'{var_}')  # 进行替换操作
            # 将用例转换回列表，并返回
            return ast.literal_eval(case_str)
        else:
            return ast.literal_eval(case_str)

    # 总断言，并将断言结果返写excel
    def all_assertion(self, excel_data, res_data):
        """
        :param excel_data: excel用例信息，单行用例
        :param res_data: 接口响应信息
        :return: 无
        """
        excel_info, excel_code = excel_data[12], excel_data[13]  # 获取excel（断言信息）和（响应码）
        assertion_result = {'res_info': True, 'res_code': True}  # 用于存储断言结果
        column_ = 15  # 定义excel中（执行结果）所在列
        result_column = 16  # 定义excel中（接口耗时）所在列

        if res_data:
            # 获取接口返回的（响应码）、（响应信息）、（接口耗时）
            res_code, res_info, res_time = res_data[0], res_data[1], res_data[2]
        else:
            res_code, res_info, res_time = "None", "None", "None"

        # 如果（响应信息）或（响应码）不为空时
        if excel_info or excel_code:

            if excel_info:  # 如果（响应信息）不为空
                assertion_result_info = Assertion().response_info_assertion(excel_info, res_info)  # （响应信息）断言
                assertion_result['res_info'] = assertion_result_info  # 将断言结果加入字典
            if excel_code:  # 如果（响应码）不为空
                assertion_result_code = Assertion().response_code_assertion(excel_code, res_code)  # （响应码）断言
                assertion_result['res_code'] = assertion_result_code  # 将断言结果加入字典

        # 获取excel路径、sheet页名称、所在行：用于反写执行结果
        path_, sheet_name, row_ = excel_data[-1][0], excel_data[-1][1], excel_data[-1][2]
        workbook = openpyxl.load_workbook(path_)  # 打开需要编辑的excel文件
        sheet = workbook[sheet_name]  # 打开需要编辑的sheet页名称

        # 将断言结果反写回excel：
        result_value = ['执行成功',
                        f'执行失败\n失败原因：（响应信息）断言失败\n实际响应信息：{res_info}',
                        f'执行失败\n失败原因：（响应码）断言失败\n实际响应状态码：{res_code}',
                        f'执行失败\n失败原因:执行失败原因：（响应码）和（响应信息）断言失败\n实际（响应码）：{res_code}\n'
                        f'实际（响应信息）：{res_info}']

        if all(assertion_result.values()):  # 当结果都为True时
            sheet.cell(column=column_, row=row_, value=result_value[0])  # 将（执行成功）返写回excel
        elif not all(assertion_result.values()):  # （响应信息）和（响应码）如果有一个失败
            if (assertion_result['res_code']) and (not assertion_result['res_info']):
                sheet.cell(column=column_, row=row_, value=result_value[1])
            elif (assertion_result['res_info']) and (not assertion_result['res_code']):
                sheet.cell(column=column_, row=row_, value=result_value[2])
            else:  # 如果res_info和res_code都能获取到，则判断（响应信息）和（响应码）都失败
                sheet.cell(column=column_, row=row_, value=result_value[3])
        elif len(assertion_result) == 0:  # 当结果不为True也不为False时，也就是（响应信息）和（响应码）都没写
            sheet.cell(column=column_, row=row_, value='执行成功')

        # 【接口耗时】
        if res_time != 'None':
            # 将接口运行耗时反写回excel
            sheet.cell(column=result_column, row=row_, value=f'{round(res_time, 1)}秒')

        try:
            # 保存excel
            workbook.save(path_)
        except PermissionError:
            raise 'excel未关闭，请关闭excel文件后执行！'

    # 【web_UI】--------------------------------------------------------------------------------------------------------
    # 打开浏览器
    def open_browser(self, case_data):
        # 如果报错， 则说明主机IP写得是实际IP，否则则说明通过配置文件获取IP
        try:
            int(case_data[7])
        except ValueError:
            ip_ = f'{case_data[7]}{case_data[8]}'
        else:
            # 读取yaml文件
            yaml_datas = YamlOperate().read_yaml(self.config_path)
            url_ = yaml_datas['URL']

            # 通过下标读取主机IP
            ip_ = url_[int(case_data[7])]
        # 打开浏览器，并输入url
        SeleniumTool().open_browser(ip_)

    # 点击
    def click(self,case_data):
        loc_mode = case_data[7]  # 获取定位方式
        ele_info = case_data[8]  # 获取元素信息
        # 点击元素
        SeleniumTool().click_ele(loc_mode=loc_mode,ele_info=ele_info)

    # 输入
    def input(self, case_data):
        loc_mode = case_data[7]  # 获取定位方式
        ele_info = case_data[8]  # 获取元素信息
        txt = case_data[9]  # 获取输入内容
        # 输入内容
        SeleniumTool().input_data(loc_mode=loc_mode,ele_info=ele_info,txt=txt)

    # 关闭浏览器
    def quit(self,case_data):
        SeleniumTool().shut_browser()  # 关闭浏览器

    # 【公用】----------------------------------------------------------------------------------------------------------
    # 等待
    def wait(self, case_data):
        # 获取等待时间
        wait_time = case_data[9]
        # 执行等待代码
        Base().wait_time(wait_time)



